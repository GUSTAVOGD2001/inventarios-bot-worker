import asyncio
import io
import logging
import time
from datetime import datetime, timedelta

import httpx
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from ..auth import require_api_key
from ..db import get_pool
from ..error_handler import log_endpoint_errors
from ..models import ShopCreate, ShopUpdate

logger = logging.getLogger(__name__)

router = APIRouter(dependencies=[Depends(require_api_key)])

DDVC_GRAPHQL_URL = "https://tiendaddvc.mx/graphql"

# ═══════════════════════════════════════════════════════════════
# SCRAPING DDVC → EXCEL
# ═══════════════════════════════════════════════════════════════

DDVC_FULL_CATALOG_QUERY = """
query GetFullCatalog($pageSize: Int!, $currentPage: Int!) {
  products(filter: {}, pageSize: $pageSize, currentPage: $currentPage) {
    items {
      sku
      name
      url_key
      is_salable
      stock_status
      type_id
      meta_title
      meta_description
      meta_keyword
      short_description { html }
      description { html }
      categories { name url_path }
      price_range {
        minimum_price {
          regular_price { value currency }
          final_price { value currency }
          discount { amount_off percent_off }
        }
        maximum_price {
          regular_price { value currency }
          final_price { value currency }
        }
      }
      media_gallery { url label position }
      image { url label }
      small_image { url label }
      thumbnail { url label }
    }
    page_info { current_page total_pages }
    total_count
  }
}
"""

DDVC_FULL_CATALOG_QUERY_FALLBACK = """
query GetFullCatalog($pageSize: Int!, $currentPage: Int!) {
  products(filter: {}, pageSize: $pageSize, currentPage: $currentPage) {
    items {
      sku
      name
      url_key
      is_salable
      type_id
      short_description { html }
      description { html }
      categories { name url_path }
      price_range {
        minimum_price {
          regular_price { value currency }
          final_price { value currency }
          discount { amount_off percent_off }
        }
      }
      media_gallery { url label position }
      image { url label }
    }
    page_info { current_page total_pages }
    total_count
  }
}
"""


async def _fetch_ddvc_full_catalog() -> list[dict]:
    """Fetch completo del catálogo DDVC con todos los campos disponibles."""
    page_size = 100
    timeout = 90.0
    sleep_seconds = 0.35
    current_page = 1
    total_pages = 1
    all_items = []
    active_query = DDVC_FULL_CATALOG_QUERY

    async with httpx.AsyncClient(
        timeout=timeout,
        headers={"User-Agent": "Agente-Portales-Libertad"},
    ) as client:
        while current_page <= total_pages:
            for attempt in range(3):
                try:
                    response = await client.post(
                        DDVC_GRAPHQL_URL,
                        json={
                            "query": active_query,
                            "variables": {"pageSize": page_size, "currentPage": current_page},
                        },
                    )
                    response.raise_for_status()
                    data = response.json()
                    if data.get("errors"):
                        error_str = str(data["errors"])
                        if any(f in error_str for f in ("stock_status", "meta_title", "meta_description", "meta_keyword")):
                            logger.warning("DDVC schema error, using fallback query")
                            active_query = DDVC_FULL_CATALOG_QUERY_FALLBACK
                            continue
                        raise RuntimeError(f"GraphQL errors: {data['errors']}")
                    break
                except Exception as exc:
                    if attempt >= 2:
                        raise
                    logger.warning("DDVC page=%s attempt=%s error=%s", current_page, attempt + 1, exc)
                    await asyncio.sleep(sleep_seconds)
            else:
                raise RuntimeError(f"Failed to fetch page {current_page}")

            products = data.get("data", {}).get("products", {})
            total_pages = products.get("page_info", {}).get("total_pages", 1)
            items = products.get("items") or []
            all_items.extend(items)

            if current_page >= total_pages:
                break
            await asyncio.sleep(sleep_seconds)
            current_page += 1

    logger.info("DDVC catalog fetched: %s items, %s pages", len(all_items), total_pages)
    return all_items


def _items_to_excel_bytes(items: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo DDVC"

    headers = [
        "SKU", "Nombre", "URL Key", "Disponible", "Stock Status", "Tipo",
        "Precio Regular", "Precio Final", "Moneda", "Descuento %", "Descuento $",
        "Categoría Principal", "Ruta Categoría",
        "Descripción Corta", "Descripción Completa",
        "Meta Title", "Meta Description", "Meta Keywords",
        "Imagen Principal URL", "Imagen Principal Label",
        "Galería URLs (|)", "Galería Labels (|)",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(items, 2):
        if not item:
            continue
        sku = (item.get("sku") or "").strip()
        if not sku:
            continue

        pr = (item.get("price_range") or {}).get("minimum_price") or {}
        reg = pr.get("regular_price") or {}
        fin = pr.get("final_price") or {}
        disc = pr.get("discount") or {}
        cats = item.get("categories") or []
        gallery = item.get("media_gallery") or []
        main_img = item.get("image") or {}

        row_data = [
            sku,
            item.get("name") or "",
            item.get("url_key") or "",
            "Sí" if item.get("is_salable") else "No",
            item.get("stock_status") or "",
            item.get("type_id") or "",
            reg.get("value"), fin.get("value"),
            reg.get("currency") or fin.get("currency") or "MXN",
            disc.get("percent_off"), disc.get("amount_off"),
            cats[0].get("name") if cats else "",
            cats[0].get("url_path") if cats else "",
            (item.get("short_description") or {}).get("html") or "",
            (item.get("description") or {}).get("html") or "",
            item.get("meta_title") or "",
            item.get("meta_description") or "",
            item.get("meta_keyword") or "",
            main_img.get("url") or "",
            main_img.get("label") or "",
            " | ".join(g.get("url", "") for g in gallery if g.get("url")),
            " | ".join(g.get("label", "") for g in gallery if g.get("label")),
        ]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    ws.column_dimensions["N"].width = 40
    ws.column_dimensions["O"].width = 40
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Fotos Editadas")
    for i, h in enumerate(["SKU", "URL Foto Original", "URL Foto Editada", "Notas"], 1):
        ws2.cell(row=1, column=i, value=h).font = Font(bold=True)
    ws2.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@router.post("/proyecto-libertad/scrape-catalog")
@log_endpoint_errors
async def scrape_ddvc_catalog():
    """Scraping completo de DDVC → Excel descargable."""
    logger.info("Starting full DDVC catalog scrape")
    start_time = time.time()
    items = await _fetch_ddvc_full_catalog()
    if not items:
        raise HTTPException(status_code=500, detail="No se obtuvieron productos de DDVC")
    excel_bytes = _items_to_excel_bytes(items)
    elapsed = time.time() - start_time
    logger.info("Excel generated: %s items, %.1fs", len(items), elapsed)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="catalogo_ddvc_{timestamp}.xlsx"'},
    )


@router.get("/proyecto-libertad/shops")
@log_endpoint_errors
async def list_shops():
    pool = await get_pool()
    rows = await pool.fetch("SELECT * FROM shops ORDER BY is_primary DESC, name")
    return {"items": [dict(r) for r in rows]}


@router.post("/proyecto-libertad/shops", status_code=201)
@log_endpoint_errors
async def create_shop(body: ShopCreate):
    pool = await get_pool()
    try:
        row = await pool.fetchrow(
            """INSERT INTO shops (name, slug, shopify_shop, shopify_client_id,
               shopify_client_secret, shopify_api_version, in_stock_qty,
               out_of_stock_qty, is_active, notes)
               VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10) RETURNING *""",
            body.name, body.slug.lower(), body.shopify_shop,
            body.shopify_client_id, body.shopify_client_secret,
            body.shopify_api_version, body.in_stock_qty,
            body.out_of_stock_qty, body.is_active, body.notes,
        )
    except Exception as e:
        if "unique" in str(e).lower():
            raise HTTPException(status_code=409, detail=f"Ya existe una tienda con slug '{body.slug}'")
        raise
    return dict(row)


@router.put("/proyecto-libertad/shops/{shop_id}")
@log_endpoint_errors
async def update_shop(shop_id: int, body: ShopUpdate):
    pool = await get_pool()
    existing = await pool.fetchrow("SELECT * FROM shops WHERE id = $1", shop_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Tienda no encontrada")
    fields, params, idx = [], [], 1
    for f in ("name", "shopify_shop", "shopify_client_id", "shopify_client_secret",
              "shopify_api_version", "in_stock_qty", "out_of_stock_qty", "is_active", "notes"):
        val = getattr(body, f)
        if val is not None:
            fields.append(f"{f} = ${idx}")
            params.append(val)
            idx += 1
    if not fields:
        return dict(existing)
    fields.append("updated_at = now()")
    params.append(shop_id)
    row = await pool.fetchrow(
        f"UPDATE shops SET {', '.join(fields)} WHERE id = ${idx} RETURNING *", *params
    )
    return dict(row)


@router.delete("/proyecto-libertad/shops/{shop_id}", status_code=204)
@log_endpoint_errors
async def delete_shop(shop_id: int):
    pool = await get_pool()
    shop = await pool.fetchrow("SELECT * FROM shops WHERE id = $1", shop_id)
    if not shop:
        raise HTTPException(status_code=404, detail="Tienda no encontrada")
    if shop["is_primary"]:
        raise HTTPException(status_code=400, detail="No se puede eliminar la tienda primaria")
    await pool.execute("DELETE FROM shops WHERE id = $1", shop_id)


@router.post("/proyecto-libertad/shops/{shop_id}/test-connection")
@log_endpoint_errors
async def test_shop_connection(shop_id: int):
    """Prueba la conexión a Shopify de una tienda."""
    pool = await get_pool()
    shop = await pool.fetchrow("SELECT * FROM shops WHERE id = $1", shop_id)
    if not shop:
        raise HTTPException(status_code=404, detail="Tienda no encontrada")
    if shop["shopify_shop"] == "placeholder":
        return {"success": False, "error": "Credenciales no configuradas (placeholder)"}
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            token_resp = await client.post(
                f"https://{shop['shopify_shop']}/admin/oauth/access_token",
                json={"client_id": shop["shopify_client_id"],
                      "client_secret": shop["shopify_client_secret"],
                      "grant_type": "client_credentials"},
            )
            token_resp.raise_for_status()
            token = token_resp.json().get("access_token")
            if not token:
                return {"success": False, "error": "No se obtuvo access_token"}
            gql_resp = await client.post(
                f"https://{shop['shopify_shop']}/admin/api/{shop['shopify_api_version']}/graphql.json",
                headers={"X-Shopify-Access-Token": token},
                json={"query": "{ shop { name } }"},
            )
            gql_resp.raise_for_status()
            shop_name = gql_resp.json().get("data", {}).get("shop", {}).get("name", "?")
            return {"success": True, "shop_name": shop_name}
    except Exception as exc:
        return {"success": False, "error": str(exc)}


async def _get_shop_token(shop: dict) -> str | None:
    """Obtiene access token de Shopify para una tienda."""
    if shop["shopify_shop"] == "placeholder":
        return None
    async with httpx.AsyncClient(timeout=15.0) as client:
        resp = await client.post(
            f"https://{shop['shopify_shop']}/admin/oauth/access_token",
            json={"client_id": shop["shopify_client_id"],
                  "client_secret": shop["shopify_client_secret"],
                  "grant_type": "client_credentials"},
        )
        resp.raise_for_status()
        return resp.json().get("access_token")


async def _fetch_shop_revenue(shop: dict, days: int) -> dict:
    """Fetch total de ventas de una tienda vía Shopify REST API."""
    result = {
        "shop_id": shop["id"], "shop_name": shop["name"],
        "slug": shop["slug"], "is_primary": shop["is_primary"],
        "total_sales": 0, "order_count": 0, "currency": "MXN",
    }
    try:
        token = await _get_shop_token(shop)
        if not token:
            result["error"] = "Credenciales no configuradas"
            return result

        since = (datetime.utcnow() - timedelta(days=days)).isoformat() + "Z"
        base_url = (
            f"https://{shop['shopify_shop']}/admin/api/{shop['shopify_api_version']}"
            f"/orders.json?status=any&created_at_min={since}"
            f"&fields=id,total_price,financial_status&limit=250"
        )

        total_sales = 0.0
        order_count = 0
        page_url = base_url

        async with httpx.AsyncClient(timeout=30.0) as client:
            while page_url:
                resp = await client.get(page_url, headers={"X-Shopify-Access-Token": token})
                resp.raise_for_status()
                orders = resp.json().get("orders", [])
                for order in orders:
                    if order.get("financial_status") in ("paid", "partially_paid", "refunded", "partially_refunded"):
                        try:
                            total_sales += float(order.get("total_price", 0))
                        except (TypeError, ValueError):
                            pass
                        order_count += 1

                link = resp.headers.get("Link", "")
                page_url = None
                if 'rel="next"' in link:
                    for part in link.split(","):
                        if 'rel="next"' in part:
                            page_url = part.split("<")[1].split(">")[0]
                            break

        result["total_sales"] = round(total_sales, 2)
        result["order_count"] = order_count
    except Exception as exc:
        logger.error("Revenue fetch failed shop=%s: %s", shop["id"], exc)
        result["error"] = str(exc)
    return result


@router.get("/proyecto-libertad/shops/{shop_id}/revenue")
@log_endpoint_errors
async def shop_revenue(shop_id: int, days: int = Query(30, ge=1, le=365)):
    pool = await get_pool()
    shop = await pool.fetchrow("SELECT * FROM shops WHERE id = $1", shop_id)
    if not shop:
        raise HTTPException(status_code=404, detail="Tienda no encontrada")
    return await _fetch_shop_revenue(dict(shop), days)


@router.get("/proyecto-libertad/revenue-summary")
@log_endpoint_errors
async def revenue_summary(days: int = Query(30, ge=1, le=365)):
    """Resumen de ventas de TODAS las tiendas activas, consultadas en paralelo."""
    pool = await get_pool()
    shops = await pool.fetch("SELECT * FROM shops WHERE is_active = true ORDER BY is_primary DESC, name")
    tasks = [_fetch_shop_revenue(dict(s), days) for s in shops]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    shop_results = []
    for r in results:
        if isinstance(r, Exception):
            shop_results.append({"error": str(r), "total_sales": 0, "order_count": 0})
        else:
            shop_results.append(r)

    grand_total = sum(r.get("total_sales", 0) for r in shop_results)
    total_orders = sum(r.get("order_count", 0) for r in shop_results)

    return {
        "period_days": days,
        "grand_total_sales": round(grand_total, 2),
        "total_orders": total_orders,
        "shops": shop_results,
    }
